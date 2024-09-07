import os
import time
import csv
from openpyxl import load_workbook
from deep_translator import GoogleTranslator
from pystyle import Write, Colors, Center, Box, System
from tqdm import tqdm

def translate_headers(headers):
    translated_headers = []
    for header in headers:
        header = header.replace('_', ' ').capitalize()
        translated_header = GoogleTranslator(source='auto', target='ru').translate(header)
        translated_headers.append(translated_header)
    return translated_headers

def search_data_in_file(file_path, search_query):
    found_data = []
    try:
        if file_path.endswith('.csv'):
            csv.field_size_limit(35 * 1024 * 1024 * 1024)
            with open(file_path, 'r', encoding='utf-8-sig') as file:
                delimiter = ';'
                first_line = file.readline().strip()
                if '\t' in first_line:
                    delimiter = '\t'
                elif '|' in first_line:
                    delimiter = '|'
                elif ',' in first_line:
                    delimiter = ','
                elif '\r' in first_line:
                    delimiter = '\r'
                elif '\n' in first_line:
                    delimiter = '\n'
                file.seek(0)
                reader = csv.reader(file, delimiter=delimiter)
                headers = next(reader)
                russian_headers = translate_headers(headers)
                pbar = tqdm(reader, desc=f"Обработка {file_path}", unit="строка", ascii=True, leave=False)
                for row in pbar:
                    full_name = ' '.join(row[2:5]).strip('\\"').lower()
                    found = False
                    if search_query.lower() in full_name:
                        found = True
                    else:
                        for value in row:
                            if search_query.lower() in value.lower():
                                found = True
                                break
                    if found:
                        found_row = {}
                        for header, value in zip(russian_headers, row):
                            found_row[header.strip('\\"')] = value.strip('\\"') if value else 'None'
                        found_data.append(found_row)
        elif file_path.endswith('.xlsx'):
            workbook = load_workbook(file_path)
            worksheet = workbook.active
            headers = [cell for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
            russian_headers = translate_headers(headers)
            pbar = tqdm(worksheet.iter_rows(min_row=2, values_only=True), desc=f"Обработка {file_path}", unit="строка", ascii=True, leave=False)
            for row in pbar:
                full_name = ' '.join(str(cell) for cell in row[2:5]).strip('\\"').lower()
                found = False
                if search_query.lower() in full_name:
                    found = True
                else:
                    for value in row:
                        if search_query.lower() in str(value).lower():
                            found = True
                            break
                if found:
                    found_row = {}
                    for header, value in zip(russian_headers, row):
                        found_row[header.strip('\\"')] = str(value).strip('\\"') if value else 'None'
                    found_data.append(found_row)
    except Exception as e:
        Write.Print(f"Ошибка при обработке файла {file_path}: {e}\n", Colors.red_to_yellow, interval=0.0001)
    return found_data

def search_data(search_query):
    start_time = time.time()
    files = [file for file in os.listdir() if file.endswith('.csv') or file.endswith('.xlsx')]
    if not files:
        Write.Print("\nНет CSV или XLSX файлов в текущей директории.\n\n", Colors.red_to_yellow, interval=0.0001)
        return
    all_found_data = {}
    total_results = 0
    for file_path in files:
        found_data = search_data_in_file(file_path, search_query)
        if found_data:
            all_found_data[file_path] = found_data
            total_results += len(found_data)
    elapsed_time = time.time() - start_time

    if elapsed_time < 60:
        time_str = f"{elapsed_time:.2f} секунд"
    else:
        minutes, seconds = divmod(elapsed_time, 60)
        time_str = f"{int(minutes)} минут {seconds:.2f} секунд"

    output_file = 'found_data.txt'
    with open(output_file, 'w', encoding='utf-8') as output:
        for file_path, data in all_found_data.items():
            Write.Print(Box.DoubleCube(f"Файл: {file_path}\nЗапрос: {search_query}\n"), Colors.red_to_yellow, interval=0.0000000000000001)
            output.write(f"Файл: {file_path}\n")
            output.write(f"Запрос: {search_query}\n\n")
            if len(data) > 100:
                Write.Print(f"Размер информации слишком большой ({len(data)} записей), смотрите файл: {output_file}\n", Colors.green_to_yellow, interval=0.001)
                for found_row in data:
                    output.write("\n┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")
                    for key, value in found_row.items():
                        output.write(f"┣ {key}: {value}\n")
                    output.write("┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n")
            else:
                for found_row in data:
                    Write.Print("\n┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n", Colors.red_to_green, interval=0.001)
                    output.write("\n┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")
                    for key, value in found_row.items():
                        Write.Print(f"┣ {key}: {value}\n", Colors.green_to_yellow, interval=0.0000000000000001)
                        output.write(f"┣ {key}: {value}\n")
                    Write.Print("┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n", Colors.red_to_green, interval=0.001)
                    output.write("┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n")

        Write.Print(f"\nВсего найдено запросов: {total_results}\n", Colors.green_to_yellow, interval=0.0000000000000001)
        Write.Print(f"Время выполнения: {time_str}\n", Colors.green_to_yellow, interval=0.001)
        Write.Print(f"\nРезультаты сохранены в файле: {output_file}\n", Colors.red_to_yellow, interval=0.0000000000000001)
        output.write(f"\nВсего найдено запросов: {total_results}\n")
        output.write(f"Время выполнения: {time_str}\n")
            
def display_banner():
    os.system('cls' if os.name == 'nt' else 'clear')
    banner = """

████████▄   ▄██████▄  ▀████    ▐████▀      ▀█████████▄     ▄████████    ▄████████    ▄████████ 
███   ▀███ ███    ███   ███▌   ████▀         ███    ███   ███    ███   ███    ███   ███    ███ 
███    ███ ███    ███    ███  ▐███           ███    ███   ███    ███   ███    █▀    ███    █▀  
███    ███ ███    ███    ▀███▄███▀          ▄███▄▄▄██▀    ███    ███   ███         ▄███▄▄▄     
███    ███ ███    ███    ████▀██▄          ▀▀███▀▀▀██▄  ▀███████████ ▀███████████ ▀▀███▀▀▀     
███    ███ ███    ███   ▐███  ▀███           ███    ██▄   ███    ███          ███   ███    █▄  
███   ▄███ ███    ███  ▄███     ███▄         ███    ███   ███    ███    ▄█    ███   ███    ███ 
████████▀   ▀██████▀  ████       ███▄      ▄█████████▀    ███    █▀   ▄████████▀    ██████████ 
                                    РАЗРАБОТЧИК: @KADICK1        
"""
    Write.Print(banner, Colors.red_to_yellow, interval=0.0001)

while True:
    display_banner()
    menu = "\n1. Пробив по базе.\n2. Выход\n"
    Write.Print(menu, Colors.red_to_yellow, interval=0.0001)
    choice = input("\nВыберите опцию: ")
    
    if choice == "1":
        os.system('cls' if os.name == 'nt' else 'clear')
        display_banner()
        search_query = input("Введите запрос: ")
        os.system('cls' if os.name == 'nt' else 'clear')
        display_banner()
        search_data(search_query)
        input("Нажмите Enter для возврата в меню...")

    elif choice == "2":
        break

    else:
        Write.Print("Неверный выбор, попробуйте снова.\n", Colors.red_to_yellow, interval=0.0001)
